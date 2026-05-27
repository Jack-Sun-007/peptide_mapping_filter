import os
import re
import warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 氨基酸序列数字补全至3位方便排序
def pad_positions(position):
    try:
        start, end = position.split('-')
        return f"{start.zfill(3)}-{end.zfill(3)}"
    except ValueError:
        # 如果无法正常分割成两部分，就返回原始字符串
        return position


# 查找文件夹的csv或excel
def get_csv(filepath):
    try:
        files = os.listdir(filepath)  # 列出目录中的所有文件和文件夹
        files = [f for f in files if os.path.isfile(os.path.join(filepath, f))]  # 过滤出所有文件
        for i, file in enumerate(files, 1):  # 用自然数序号展示
            print(f'{i}. {file}')
        index = int(input("请输入文件的序号: "))
        file = files[index - 1]
        file_path = os.path.abspath(os.path.join(filepath, file))
        return file_path
    except FileNotFoundError:
        print("ERROR")
        return []


# 对csv进行数据清洗
def csv_dataclean(path):
    df = pd.read_csv(path)
    # 指定删除包含列'Identification'空值的行
    df = df.dropna(subset=['Identification'])
    # 优化单个样品的情况:单个样品列数为23，两个样品列数为54
    if len(df.columns) < 30:
        df['Max MS Area'] = df['MS Area'].copy()
        df['MS Area:1'] = df['MS Area'].copy()
        df['∆ ppm:1'] = df['∆ ppm'].copy()
        df['ID Type:1'] = df['ID Type'].copy()
        df['Mono Mass Exp.:1'] = df['Mono Mass Exp.'].copy()
    else:
        pass
    # 过滤mass error ±10 ppm
    df = df.loc[df['∆ ppm'] >= -10]
    df = df.loc[df['∆ ppm'] <= 10]
    # 先按轻重链排序，再按氨基酸序列排序
    df['Positions'] = df['Positions'].apply(pad_positions)
    df = df.sort_values(by=["Protein", "Positions", "Max MS Area"], ascending=[False, True, False])
    # 删掉多余列
    columns_to_drop = list()
    sample_number = 1
    for column in df.columns:
        if 'Ratio' in column:
            sample_number += 1
            columns_to_drop.append(column)
        if '%CV' in column:
            columns_to_drop.append(column)
        if 'Cond.' in column:
            columns_to_drop.append(column)
        if 'Avg' in column:
            columns_to_drop.append(column)
        if 'RT' in column and '.raw' in column:
            columns_to_drop.append(column)
        if 'M/Z' in column and '.raw' in column:
            columns_to_drop.append(column)
    columns_to_drop.extend(['Level', '∆ ppm', 'Conf. Score', 'Integration Type', 'Protein', 'Best ASR'])
    df.drop(columns=columns_to_drop, inplace=True)
    # 重新排序列
    columns = df.columns.tolist()
    msarea, deltappm, idtype, monomass, none = list(), list(), list(), list(), list()
    for i in columns:
        if 'MS Area:' in i:
            msarea.append(i)
        elif '∆ ppm:' in i:
            deltappm.append(i)
        elif 'ID Type:' in i:
            idtype.append(i)
        elif 'Mono Mass Exp.:' in i:
            monomass.append(i)
        else:
            none.append(i)
    monoppm_list = [item for pair in zip(monomass, deltappm) for item in pair]
    new_columns = none + msarea + monoppm_list + idtype
    df = df.reindex(columns=new_columns)
    # 过滤MS Area<1E4的nonspecific
    df = df[~((df['Mod'].str.contains("nonspecific", na=False)) & (df['Max MS Area'] < 10000))]
    # 过滤杂质修饰
    df = df[~df['Mod'].str.contains("Al3\\+|Ca\\+|Ca2\\+|Fe2\\+|Fe3\\+|Na\\+|2x|GasPhase", na=False)]
    # 筛选Missed Cleavages < 4
    df = df.loc[df['Missed Cleavages'] < 4]
    # 字体字号设置
    styled_df = df.style.set_properties(**{
        'font-family': 'Times New Roman',
        'font-size': '9pt'
    })
    # 导出格式化后的raw表
    filename = "1.Raw_" + str(path.split("\\")[-1].split(".")[0]) + ".xlsx"
    styled_df.to_excel(filename, engine='openpyxl', index=False)
    return df, filename[6:]


# 假阳性氧化肽段筛选并去除（依据保留时间）
def clear_fake_oxidation(df):
    # 确保 RT 是数值
    df["RT"] = pd.to_numeric(df["RT"], errors="coerce")
    # 标记是否为氧化
    df["is_oxidation"] = df["Mod_x"].str.contains("Oxidation", na=False)
    # 标记是否为无修饰肽段（Mod_x为空或含Carbamidomethylation）
    df["is_unmodified"] = (df["Mod_x"].isna() | (df["Mod_x"].str.strip() == "") | (df["Mod_x"].str.fullmatch(r"\(Carbamidomethylation\)", na=False)))
    # 分组键（可以根据需要增减）
    group_cols = ["Peptide Sequence", "Charge St."]
    # 计算每组无修饰肽段的 RT（取最小值，最保守）
    unmod_rt = (
        df[df["is_unmodified"]]
            .groupby(group_cols)["RT"]
            .min()
            .rename("RT_unmod")
    )
    # 合并回原表
    df = df.merge(unmod_rt, on=group_cols, how="left")
    # 判定假阳性氧化（氧化肽段RT>None肽段RT-0.2）
    df["false_positive_ox"] = (
            df["is_oxidation"] &
            df["RT_unmod"].notna() &
            (df["RT"] >= df["RT_unmod"] - 0.2)
    )
    # 过滤掉假阳性氧化
    df_filtered = df[~df["false_positive_ox"]].copy()
    return df_filtered


# 假阳性肽段筛选并去除（依据保留时间）
def _dedup_one_pep(g: pd.DataFrame) -> pd.DataFrame:
    # 电荷=1：同电荷重复只保留最前面
    if g["Charge St."].nunique(dropna=False) == 1:
        return g.loc[~g.duplicated(subset=["Charge St."], keep="first")]
    # 电荷>=2：先用“Max MS Area”最大的那行RT作为标准RT
    rt_std = g.loc[g["Max MS Area"].idxmax(), "RT"]
    # 对每个电荷价态，保留RT距离标准RT最近的那行（若并列，保留原表中更靠前的）
    g2 = g.assign(_rt_dist=(g["RT"] - rt_std).abs(), _ord=range(len(g)))
    keep_idx = (
        g2.sort_values(["Charge St.", "_rt_dist", "_ord"], ascending=[True, True, True])
          .groupby("Charge St.", sort=False)
          .head(1)
          .index
    )
    return g.loc[keep_idx].sort_index()


# 对PTM模板进行数据合并
def merge_data(Template_path, df, df_filename):
    a_df = pd.read_excel(Template_path)
    b_df = df
    # 合并数据，依据理论分子量和序列进行合并
    merged_df = b_df.merge(a_df, on=["Mono Mass Theo.", "Peptide Sequence"], how="inner")
    # 取修饰和带电荷数的交集（每个 Peptide + Mod_x 的 Charge 集合）
    charge_sets = (
        merged_df.groupby(['Peptide Sequence', 'Mod_x'], dropna=False)['Charge St.']
            .apply(lambda x: set(x.dropna()))
    )
    # 对同一 Peptide，不同 Mod_x 的 Charge 取交集
    common_charges = (
        charge_sets
            .groupby(level=0)
            .agg(lambda sets: set.intersection(*sets) if len(sets) > 1 else next(iter(sets)))
    )
    allowed = (
        common_charges
            .explode()
            .reset_index()
            .rename(columns={0: 'Charge St.'})
    )
    # 只保留 Charge St. 在交集中的行
    result_df = (
        merged_df.merge(
            allowed,
            on=['Peptide Sequence', 'Charge St.'],
            how='inner'
        )
    )
    # 将K loss肽段放到最后一行
    result_df = result_df.sort_values(
        by='Peptide Sequence',
        key=lambda s: s.str.contains('SLSL', na=False),
        kind='stable'
    ).reset_index(drop=True)
    # 去除假阳性氧化肽段：
    result_df = clear_fake_oxidation(result_df)
    # 去除假阳性None肽段：
    mask_unmod = result_df["is_unmodified"].eq(True)
    dup_in_unmod = result_df.loc[mask_unmod].duplicated(subset=["Peptide Sequence", "Charge St."], keep="first")
    result_df = result_df.drop(result_df.loc[mask_unmod].index[dup_in_unmod]).copy()
    # 去除N-端假阳性修饰：
    # Glu_Pyro_Glu
    has_mod = result_df["Mod_x"].astype(str).str.contains("Glu_Pyro_Glu", na=False).any()
    if has_mod:
        mask = result_df["Mod_x"].eq("Glu_Pyro_Glu")
        keep_mod = (result_df.loc[mask]
                    .groupby(["Peptide Sequence", "Mod_x"], group_keys=False, sort=False)
                    .apply(_dedup_one_pep))
        if mask.any():
            parts = [result_df.loc[~mask], keep_mod]
            parts = [x for x in parts if len(x) > 0]
            result_df = pd.concat(parts, axis=0).sort_index().reset_index(drop=True)
        else:
            result_df = result_df.reset_index(drop=True)
    # Gln->Pyro-Glu
    has_mod2 = result_df["Mod_x"].astype(str).str.contains("Gln->Pyro-Glu", na=False).any()
    if has_mod2:
        mask = result_df["Mod_x"].eq("Gln->Pyro-Glu")
        keep_mod = (result_df.loc[mask]
                    .groupby(["Peptide Sequence", "Mod_x"], group_keys=False, sort=False)
                    .apply(_dedup_one_pep))
        if mask.any():
            parts = [result_df.loc[~mask], keep_mod]
            parts = [x for x in parts if len(x) > 0]
            result_df = pd.concat(parts, axis=0).sort_index().reset_index(drop=True)
        else:
            result_df = result_df.reset_index(drop=True)
    #result_df.to_excel("TEST_result_df_TEST.xlsx", index=False)
    return result_df


# PTM计算，基于chatgpt5.2辅助生成，可能会有bug需要修改
def PTM_calc(df, df_filename):
    # 查找各个样品的MS Area列
    area_cols = [c for c in df.columns if c.startswith("MS Area:")]
    # 标记明细行
    df["row_type"] = "detail"
    df["chain"] = df['Identification'].str.split(':').str[0]
    # 肽段分组键
    peptide_keys = ["chain", "Positions", "Peptide Sequence"]
    result_blocks = []
    # 按肽段分组
    for _, pep_df in df.groupby(peptide_keys, dropna=False):
        # 保留原始明细
        result_blocks.append(pep_df)
        # 如果 Peptide Sequence 含有 SLSLS，则不独立计算比例
        peptide_seq = pep_df["Peptide Sequence"].iloc[0]
        if pd.notna(peptide_seq) and "SLSLS" in peptide_seq:
            continue  # 跳过比例计算
        # 每个样品的肽段总 MS Area（分母）
        total_by_sample = pep_df[area_cols].sum(axis=0)
        # 按 Mod_x + Site 分组
        for (mod, site), mod_df in pep_df.groupby(["Mod_x", "Site"], dropna=False):
            if pd.isna(mod) or mod == "(Carbamidomethylation)":
                continue  # 未修饰不生成比例行
            summary = {c: np.nan for c in df.columns}
            # 写回肽段信息
            summary["Mod_x"] = mod
            summary["Site"] = site
            summary["row_type"] = "PTM_ratio"
            # 每个样品独立计算比例
            for col in area_cols:
                denom = total_by_sample[col]
                if denom > 0:
                    summary[col] = mod_df[col].sum() / denom
                else:
                    summary[col] = np.nan
            result_blocks.append(pd.DataFrame([summary]))
    # 合并并导出
    final_df = pd.concat(result_blocks, ignore_index=True)
    # SLSLSP单独计算公式，抓取所有含 SLSLS 的真实肽段序列（去重）
    sls_df = (
        df.loc[
            df["Peptide Sequence"].notna()
            & df["Peptide Sequence"].str.contains("SLSLS"),
            ["Peptide Sequence"]
        ]
            .drop_duplicates()
            .copy()
    )
    sls_df2 = df.loc[df["Peptide Sequence"].notna() & df["Peptide Sequence"].str.contains("SLSLS")].copy()
    ms_area_cols = [c for c in sls_df2.columns if c.startswith("MS Area:")]
    # 新增“电荷校正后”的 MS Area 列，即2价态MS Area除以2，1价态不变
    for col in ms_area_cols:
        sls_df2[f"Adj {col}"] = np.where(
            sls_df2["Charge St."] == 2,
            sls_df2[col] / 2,
            sls_df2[col]
        )
    adj_cols = [c for c in sls_df2.columns if c.startswith("Adj MS Area:")]
    peptide_sum = (sls_df2.groupby("Peptide Sequence")[adj_cols].sum())
    peptide_ratio = peptide_sum.div(peptide_sum.sum())
    peptide_ratio = peptide_ratio.reset_index()
    # 按肽段长度排序
    sls_df["pep_len"] = sls_df["Peptide Sequence"].str.len()
    sls_df = sls_df.sort_values("pep_len").reset_index(drop=True)
    # 自动标记 short / middle / long
    n = len(sls_df)
    if n == 1:
        labels = ["single"]
    elif n == 2:
        labels = ["short", "long"]
    elif n == 3:
        labels = ["short", "middle", "long"]
    else:
        # 如果将来 >3 条，自动编号，避免程序崩
        labels = [f"SLSLS_{i + 1}" for i in range(n)]
    sls_df["SLSLS_type"] = labels
    # 构造 “占位比例行”
    sls_merged = pd.merge(
        peptide_ratio,
        sls_df,
        on="Peptide Sequence",
        how="left"
    )
    placeholder_rows = []
    for _, r in sls_merged.iterrows():
        row = {c: np.nan for c in final_df.columns}
        for i, j in zip(ms_area_cols, adj_cols):
            row[i] = r[j]
        row["Peptide Sequence"] = r["Peptide Sequence"]  # 真实序列
        row["row_type"] = "PTM_ratio_SLSLS"
        row["SLSLS_type"] = r["SLSLS_type"]
        placeholder_rows.append(pd.DataFrame([row]))
    # 拼接到最终 dataframe 最后
    final_df = pd.concat(
        [final_df] + placeholder_rows,
        ignore_index=True
    )
    # 合并计算脱酰胺和异构化的比例
    inherit_cols = [
        "Peptide Sequence",
        "chain"
    ]
    final_df[inherit_cols] = final_df[inherit_cols].ffill()
    # 仅保留 PTM_ratio 行
    ratio_df = final_df[final_df["row_type"].isin(["PTM_ratio", "PTM_ratio_SLSLS"])].copy()
    # 找MS Area列
    ratio_cols = [c for c in ratio_df.columns if c.startswith("MS Area:")]
    # 解析修饰位点函数
    def parse_site(site):
        """
        输入: 'N138', '~N139', 'D402'
        输出: ('N', 138)
        """
        if pd.isna(site):
            return None, None
        m = re.search(r"[A-Z]\d+", str(site))
        if m:
            s = m.group()
            return s[0], int(s[1:])
        return None, None
    ratio_df[["AA", "POS"]] = ratio_df["Site"].apply(lambda x: pd.Series(parse_site(x)))
    # 定义合并脱酰胺和异构化的规则，"pos_diff"为合并位点计算，目前设置相差2个位点就合并
    rules = [
        {
            "name": "Deamidation_SUM",
            "mods": ["Deamidation", "Deamidation Succinimide", "Deamidation, (Carbamidomethylation)", "Deamidation Succinimide, (Carbamidomethylation)"],
            "aa": "N",
            "pos_diff": 2
        },
        {
            "name": "Isomerization_SUM",
            "mods": ["Asp Succinimide", "Isomerization", "Asp Succinimide, (Carbamidomethylation)", "Isomerization, (Carbamidomethylation)"],
            "aa": "D",
            "pos_diff": 2
        }
    ]
    new_rows = []
    # 按肽段序列进行脱酰胺和异构化的合并计算
    for pep, pep_df in ratio_df.groupby("Peptide Sequence"):
        for rule in rules:
            sub = pep_df[pep_df["Mod_x"].isin(rule["mods"])].copy()
            if sub.empty:
                continue
            # 只保留目标氨基酸
            sub = sub[sub["AA"] == rule["aa"]]
            if sub.empty:
                continue
            # 按位点聚类
            used = set()
            for i, row in sub.iterrows():
                if i in used:
                    continue
                cluster = sub[
                    (sub["POS"] - row["POS"]).abs() <= rule["pos_diff"]
                    ]
                used.update(cluster.index)
                if len(cluster) < 2:
                    continue
                # 数值列加合
                summed = cluster[ratio_cols].sum()
                new_row = {
                    "Peptide Sequence": pep,
                    "Mod_x": rule["name"],
                    "Site": f"{rule['aa']}{row['POS']}±{rule['pos_diff']}",
                    "row_type": row["row_type"]
                }

                for col in ratio_cols:
                    new_row[col] = summed[col]
                new_rows.append(new_row)
    # 合并回原数据表中
    new_df = pd.DataFrame(new_rows)
    final_df2 = pd.concat([final_df, new_df], ignore_index=True)
    # 合并的修饰移动至各自肽段中
    final_df3 = final_df2[final_df2["row_type"] != "PTM_ratio_SLSLS"]
    slsl_df = final_df2[final_df2["row_type"] == "PTM_ratio_SLSLS"]
    move_mods = {"Deamidation_SUM", "Isomerization_SUM"}
    result_groups = []
    for peptide, group in final_df3.groupby("Peptide Sequence", sort=False):
        # 含有SLSLSP序列完全不移动
        if isinstance(peptide, str) and "SLSLS" in peptide:
            result_groups.append(group)
            continue
        # 其他序列：SUM行移到各自肽段的最后
        normal_rows = group[~group["Mod_x"].isin(move_mods)]
        moved_rows = group[group["Mod_x"].isin(move_mods)]
        # 先弄普通行，再弄SUM行
        new_group = pd.concat([normal_rows, moved_rows], ignore_index=True)
        result_groups.append(new_group)
    # 保存结果
    result_df2 = pd.concat(result_groups, ignore_index=True)
    result_df3 = pd.concat([result_df2, slsl_df], ignore_index=True)
    # 生成终表并进行字体格式设置
    style_df = result_df3.style.set_properties(**{
        'font-family': 'Times New Roman',
        'font-size': '9pt'
    })
    numeric_cols = result_df3.select_dtypes(include=[np.number]).columns.tolist()
    # 导出文件命名
    filename = "2.PTM_" + df_filename
    style_df.to_excel(filename, index=False)
    # 百分比和底色格式设置
    wb = load_workbook(filename)
    ws = wb.active
    # 表头-列号映射
    col_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    row_type_col = col_index["row_type"]
    for row in range(2, ws.max_row + 1):
        # 只处理 row_type == PTM_ratio 的行
        if ws.cell(row=row, column=row_type_col).value in ("PTM_ratio", "PTM_ratio_SLSLS"):
            for col in numeric_cols:
                c = ws.cell(row=row, column=col_index[col])
                if isinstance(c.value, (int, float)):
                    # Excel 数值型百分比，1 位小数，底色天青色
                    c.number_format = "0.0%"
                    c.fill = PatternFill("solid", fgColor="DAEEF3")
    wb.save(filename)


def run_enzyme_process(template_path, folder):
    data_path = get_csv(folder)
    #data_path = r"D:\PY\BioPharmaFinder_excel\PTM\2.Raw data with Trypsin\HLX13_PM_Trypsin_20260108_Stability.csv"  #测试用要删除======================
    df, filename = csv_dataclean(data_path)
    result_df = merge_data(template_path, df, filename)
    PTM_calc(result_df, filename)
    return []


def main():
    print("选择PTM模板")
    template_path = get_csv(".\\1.Template")
    #template_path = r"D:\PY\BioPharmaFinder_excel\PTM\1.Template\HLX13_PTM模板1.xlsx"  #测试用要删除======================
    print("选择原始csv文件")
    folder = ".\\2.Raw data with Trypsin"
    run_enzyme_process(template_path, folder)


if __name__ == '__main__':
    try:
        # 对FutureWarning的提示进行忽略，主要是对.apply(_dedup_one_pep))的代码忽略提示
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", FutureWarning)
            main()
    except Exception as e:
        print("出错啦！需重新运行，出错信息如下：")
        print(str(e))
        print("按任意键退出......")
        input()
